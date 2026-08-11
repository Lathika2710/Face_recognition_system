"""
app.py
AI Face Recognition & Smart Attendance System — Flask application.

Run with:
    python app.py
Then open http://127.0.0.1:5000 in your browser.

Demo login: admin / admin123  (see README.md)
"""

import os
import io
import csv
import time
import base64
from datetime import datetime, timedelta
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, flash, send_file, Response
)
from werkzeug.security import check_password_hash, generate_password_hash

import database as db
import face_utils as fu

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FACE_DATA_DIR = os.path.join(BASE_DIR, "static", "face_data")
UNKNOWN_DIR = os.path.join(FACE_DATA_DIR, "unknown")
PROFILE_DIR = os.path.join(FACE_DATA_DIR, "profiles")
os.makedirs(UNKNOWN_DIR, exist_ok=True)
os.makedirs(PROFILE_DIR, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key")

# In-memory cooldown tracker for unknown-face de-duplication: {rounded_encoding_hash: last_seen_ts}
_unknown_cooldown_cache = {}
# In-memory temp store for in-progress registration face samples: {temp_id: [ {encoding, quality, image_path} ]}
_pending_registrations = {}


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            if request.path.startswith("/api/"):
                return jsonify({"success": False, "error": "Not authenticated"}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


@app.context_processor
def inject_globals():
    return {
        "admin_name": session.get("full_name", "Admin"),
        "now": datetime.now(),
    }


def current_user_id():
    return session.get("user_id")


def get_owned_person(conn, person_id):
    return conn.execute(
        "SELECT * FROM persons WHERE id = ? AND user_id = ?",
        (person_id, current_user_id())
    ).fetchone()


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html")

    data = request.get_json(silent=True) or request.form
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""

    conn = db.get_db()
    user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()

    if user and check_password_hash(user["password_hash"], password):
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["full_name"] = user["full_name"] or user["username"]
        return jsonify({"success": True, "redirect": url_for("dashboard")})

    return jsonify({"success": False, "error": "Invalid username or password"}), 401


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "GET":
        return render_template("signup.html")

    data = request.get_json(silent=True) or request.form
    full_name = (data.get("full_name") or "").strip()
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    confirm_password = data.get("confirm_password") or ""

    if not username or not password or not confirm_password:
        return jsonify({"success": False, "error": "Username and password are required."}), 400
    if password != confirm_password:
        return jsonify({"success": False, "error": "Passwords do not match."}), 400

    conn = db.get_db()
    existing = conn.execute("SELECT 1 FROM users WHERE username = ?", (username,)).fetchone()
    if existing:
        conn.close()
        return jsonify({"success": False, "error": "That username is already taken."}), 409

    password_hash = generate_password_hash(password)
    conn.execute(
        "INSERT INTO users (username, password_hash, full_name) VALUES (?, ?, ?)",
        (username, password_hash, full_name or username)
    )
    conn.commit()
    conn.close()

    return jsonify({"success": True, "redirect": url_for("login")})


# ---------------------------------------------------------------------------
# Page routes
# ---------------------------------------------------------------------------
@app.route("/")
def home():
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html", active="dashboard")


@app.route("/recognition")
@login_required
def recognition():
    threshold = db.get_setting("recognition_threshold", "0.55")
    return render_template("recognition.html", active="recognition", threshold=threshold)


@app.route("/register")
@login_required
def register():
    unknown_id = request.args.get("unknown_id", type=int)
    unknown_snapshot = None
    if unknown_id:
        conn = db.get_db()
        row = conn.execute(
            "SELECT snapshot_path FROM unknown_faces WHERE id = ? AND resolved = 0 AND user_id = ?",
            (unknown_id, current_user_id())
        ).fetchone()
        conn.close()
        if row and row["snapshot_path"]:
            unknown_snapshot = row["snapshot_path"].replace("\\", "/")
    return render_template("register.html", active="register", unknown_id=unknown_id, unknown_snapshot=unknown_snapshot)


@app.route("/people")
@login_required
def people():
    return render_template("people.html", active="people")


@app.route("/people/<int:person_id>")
@login_required
def person_details(person_id):
    conn = db.get_db()
    person = get_owned_person(conn, person_id)
    conn.close()
    if not person:
        flash("Person not found", "error")
        return redirect(url_for("people"))
    return render_template("person_details.html", active="people", person=dict(person), person_id=person_id)


@app.route("/attendance")
@login_required
def attendance():
    return render_template("attendance.html", active="attendance")


@app.route("/history")
@login_required
def history():
    return render_template("history.html", active="history")


@app.route("/unknown-faces")
@login_required
def unknown_faces_page():
    return render_template("unknown_faces.html", active="unknown")


@app.route("/reports")
@login_required
def reports():
    return render_template("reports.html", active="reports")


@app.route("/settings")
@login_required
def settings_page():
    conn = db.get_db()
    rows = conn.execute("SELECT key, value FROM settings").fetchall()
    conn.close()
    settings = {r["key"]: r["value"] for r in rows}
    return render_template("settings.html", active="settings", settings=settings)


# ---------------------------------------------------------------------------
# API — Dashboard stats
# ---------------------------------------------------------------------------
@app.route("/api/stats")
@login_required
def api_stats():
    conn = db.get_db()
    today = db.today_str()
    user_id = current_user_id()

    total_people = conn.execute(
        "SELECT COUNT(*) c FROM persons WHERE status='Active' AND user_id = ?", (user_id,)
    ).fetchone()["c"]
    recognized_today = conn.execute(
        "SELECT COUNT(DISTINCT person_id) c FROM recognition_history "
        "WHERE user_id = ? AND is_known = 1 AND date(recognized_at) = ?", (user_id, today)
    ).fetchone()["c"]
    unknown_today = conn.execute(
        "SELECT COUNT(*) c FROM unknown_faces WHERE user_id = ? AND date(first_seen) = ?", (user_id, today)
    ).fetchone()["c"]
    attendance_today = conn.execute(
        "SELECT COUNT(*) c FROM attendance a JOIN persons p ON p.id = a.person_id "
        "WHERE p.user_id = ? AND a.date = ?", (user_id, today)
    ).fetchone()["c"]

    # people registered this month, for the "+N this month" delta
    month_start = datetime.now().strftime("%Y-%m-01")
    new_this_month = conn.execute(
        "SELECT COUNT(*) c FROM persons WHERE user_id = ? AND date(created_at) >= ?", (user_id, month_start)
    ).fetchone()["c"]

    conn.close()
    return jsonify({
        "success": True,
        "total_people": total_people,
        "recognized_today": recognized_today,
        "unknown_today": unknown_today,
        "attendance_today": attendance_today,
        "new_this_month": new_this_month,
        "engine_available": fu.FACE_RECOGNITION_AVAILABLE,
    })


@app.route("/api/system-status")
@login_required
def api_system_status():
    conn = db.get_db()
    try:
        conn.execute("SELECT 1")
        db_ok = True
    except Exception:
        db_ok = False
    conn.close()
    return jsonify({
        "success": True,
        "database_connected": db_ok,
        "ai_model_ready": fu.FACE_RECOGNITION_AVAILABLE,
    })


# ---------------------------------------------------------------------------
# API — Registration flow
# ---------------------------------------------------------------------------
@app.route("/api/register/capture", methods=["POST"])
@login_required
def api_register_capture():
    """
    Receives one webcam frame (base64), runs face detection, returns a
    quality score and box so the frontend can guide the user. Also stashes
    the encoding server-side under a temp_id for the final save step.
    """
    payload = request.get_json(force=True)
    temp_id = payload.get("temp_id")
    image_data = payload.get("image")

    if not temp_id or not image_data:
        return jsonify({"success": False, "error": "Missing temp_id or image"}), 400

    try:
        img = fu.decode_base64_image(image_data)
    except Exception:
        return jsonify({"success": False, "error": "Could not decode image"}), 400

    if img is None:
        return jsonify({"success": False, "error": "Could not decode image"}), 400

    try:
        faces = fu.detect_faces_and_encodings(img)
    except fu.FaceEngineUnavailable as e:
        return jsonify({"success": False, "error": str(e)}), 503

    if len(faces) == 0:
        return jsonify({"success": False, "error": "No face detected. Please look straight at the camera."})
    if len(faces) > 1:
        return jsonify({"success": False, "error": "Multiple faces detected. Only one person should be in frame."})

    face = faces[0]
    quality = fu.face_quality_score(img, face["box"])

    if quality < 35:
        return jsonify({
            "success": False,
            "error": "Face quality too low. Improve lighting and face the camera directly.",
            "quality": quality,
        })

    _pending_registrations.setdefault(temp_id, [])
    profile_snap = fu.save_snapshot(img, PROFILE_DIR, prefix=f"tmp_{temp_id}")
    _pending_registrations[temp_id].append({
        "encoding": face["encoding"],
        "quality": quality,
        "image_path": profile_snap,
    })

    return jsonify({
        "success": True,
        "quality": quality,
        "sample_count": len(_pending_registrations[temp_id]),
        "box": face["box"],
    })


@app.route("/api/register/save", methods=["POST"])
@login_required
def api_register_save():
    payload = request.get_json(force=True)
    temp_id = payload.get("temp_id")
    samples = _pending_registrations.get(temp_id, [])

    if len(samples) == 0:
        return jsonify({"success": False, "error": "No face samples captured. Please complete face capture first."}), 400

    full_name = (payload.get("full_name") or "").strip()
    if not full_name:
        return jsonify({"success": False, "error": "Full name is required."}), 400

    if not payload.get("consent"):
        return jsonify({"success": False, "error": "Consent checkbox must be confirmed before saving."}), 400

    conn = db.get_db()
    cur = conn.cursor()

    # Use the best-quality sample's image as the profile photo
    best_sample = max(samples, key=lambda s: s["quality"])
    rel_photo_path = os.path.relpath(best_sample["image_path"], BASE_DIR).replace("\\", "/")

    try:
        cur.execute("""
            INSERT INTO persons (user_id, full_name, person_code, age, gender, phone, email,
                                  department, role, address, dob, photo_path, status, consent_given)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Active', 1)
        """, (
            current_user_id(),
            full_name,
            payload.get("person_code") or None,
            payload.get("age") or None,
            payload.get("gender") or None,
            payload.get("phone") or None,
            payload.get("email") or None,
            payload.get("department") or None,
            payload.get("role") or None,
            payload.get("address") or None,
            payload.get("dob") or None,
            rel_photo_path,
        ))
    except Exception as e:
        conn.close()
        return jsonify({"success": False, "error": f"Could not save person: {e}"}), 400

    person_id = cur.lastrowid

    for s in samples:
        cur.execute(
            "INSERT INTO face_encodings (person_id, encoding, quality) VALUES (?, ?, ?)",
            (person_id, fu.encoding_to_blob(s["encoding"]), s["quality"])
        )

    if payload.get("unknown_id"):
        try:
            cur.execute(
                "UPDATE unknown_faces SET resolved = 1 WHERE id = ? AND user_id = ?",
                (int(payload["unknown_id"]), current_user_id())
            )
        except Exception:
            pass

    conn.commit()
    conn.close()

    del _pending_registrations[temp_id]

    return jsonify({
        "success": True,
        "person_id": person_id,
        "full_name": full_name,
        "sample_count": len(samples),
    })


@app.route("/api/register/reset", methods=["POST"])
@login_required
def api_register_reset():
    payload = request.get_json(force=True)
    temp_id = payload.get("temp_id")
    _pending_registrations.pop(temp_id, None)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# API — Live recognition
# ---------------------------------------------------------------------------
def _load_known_encodings(conn):
    rows = conn.execute("""
        SELECT fe.encoding, p.id as person_id, p.full_name
        FROM face_encodings fe
        JOIN persons p ON p.id = fe.person_id
        WHERE p.status = 'Active' AND p.user_id = ?
    """, (current_user_id(),)).fetchall()
    return [(r["person_id"], r["full_name"], fu.blob_to_encoding(r["encoding"])) for r in rows]


@app.route("/api/recognize", methods=["POST"])
@login_required
def api_recognize():
    payload = request.get_json(force=True)
    image_data = payload.get("image")
    if not image_data:
        return jsonify({"success": False, "error": "Missing image"}), 400

    try:
        img = fu.decode_base64_image(image_data)
    except Exception:
        return jsonify({"success": False, "error": "Could not decode image"}), 400
    if img is None:
        return jsonify({"success": False, "error": "Could not decode image"}), 400

    small, scale = fu.resize_for_processing(img, max_width=480)

    try:
        faces = fu.detect_faces_and_encodings(small)
    except fu.FaceEngineUnavailable as e:
        return jsonify({"success": False, "error": str(e)}), 503

    threshold_pct = float(db.get_setting("recognition_threshold", "0.55")) * 100

    conn = db.get_db()
    known = _load_known_encodings(conn)

    results = []
    for face in faces:
        person_id, name, confidence, distance = fu.compare_encoding_to_known(face["encoding"], known)
        box = [int(v / scale) for v in face["box"]]  # scale back to original frame coords

        if person_id is not None and confidence >= threshold_pct:
            # Known face — log recognition, mark attendance
            _log_recognition(conn, person_id, name, confidence, is_known=1)
            _mark_attendance(conn, person_id)
            _update_last_seen(conn, person_id)
            results.append({
                "status": "known",
                "person_id": person_id,
                "name": name,
                "confidence": confidence,
                "box": box,
            })
        else:
            # Unknown face — cooldown-guarded logging
            unknown_id = _handle_unknown_face(conn, img, face, confidence)
            results.append({
                "status": "unknown",
                "confidence": confidence,
                "box": box,
                "logged": bool(unknown_id),
                "unknown_id": unknown_id,
            })

    conn.commit()
    conn.close()

    return jsonify({"success": True, "faces": results, "count": len(results)})


def _log_recognition(conn, person_id, name, confidence, is_known):
    conn.execute(
        "INSERT INTO recognition_history (user_id, person_id, person_name, confidence, is_known) VALUES (?, ?, ?, ?, ?)",
        (current_user_id(), person_id, name, confidence, is_known)
    )


def _mark_attendance(conn, person_id):
    today = db.today_str()
    now = db.now_time_str()
    existing = conn.execute(
        "SELECT * FROM attendance WHERE person_id = ? AND date = ?", (person_id, today)
    ).fetchone()

    if existing is None:
        # First sighting today = entry
        status = "Late" if now > "09:30:00" else "Present"
        conn.execute(
            "INSERT INTO attendance (person_id, date, entry_time, status) VALUES (?, ?, ?, ?)",
            (person_id, today, now, status)
        )
    else:
        # Subsequent sighting = update exit time (rolling last-seen exit)
        conn.execute(
            "UPDATE attendance SET exit_time = ? WHERE id = ?", (now, existing["id"])
        )


def _update_last_seen(conn, person_id):
    conn.execute(
        "UPDATE persons SET last_seen = ? WHERE id = ?",
        (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), person_id)
    )


def _handle_unknown_face(conn, full_frame_img, face, confidence):
    """De-duplicate unknown faces using an in-memory cooldown keyed by a
    coarse rounding of the encoding vector (a cheap approximate-match)."""
    cooldown_seconds = int(db.get_setting("unknown_cooldown_seconds", "30"))
    key = tuple(round(v, 1) for v in face["encoding"][:8])  # coarse fingerprint
    now_ts = time.time()

    last_ts = _unknown_cooldown_cache.get(key)
    if last_ts and (now_ts - last_ts) < cooldown_seconds:
        # still within cooldown — try to bump the matching DB record's count instead of spamming
        recent = conn.execute(
            "SELECT * FROM unknown_faces WHERE resolved = 0 AND user_id = ? ORDER BY last_seen DESC LIMIT 20",
            (current_user_id(),)
        ).fetchall()
        if recent:
            row = recent[0]
            conn.execute(
                "UPDATE unknown_faces SET detection_count = detection_count + 1, last_seen = ? WHERE id = ?",
                (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), row["id"])
            )
            return row["id"]
        return None

    _unknown_cooldown_cache[key] = now_ts
    snap_path = fu.save_snapshot(full_frame_img, UNKNOWN_DIR, prefix="unknown")
    rel_path = os.path.relpath(snap_path, BASE_DIR).replace("\\", "/")

    cursor = conn.execute(
        "INSERT INTO unknown_faces (user_id, snapshot_path, confidence, encoding) VALUES (?, ?, ?, ?)",
        (current_user_id(), rel_path, confidence, fu.encoding_to_blob(face["encoding"]))
    )
    unknown_id = cursor.lastrowid
    _log_recognition(conn, None, "Unknown", confidence, is_known=0)
    return unknown_id


# ---------------------------------------------------------------------------
# API — People management
# ---------------------------------------------------------------------------
@app.route("/api/people")
@login_required
def api_people_list():
    q = request.args.get("q", "").strip()
    department = request.args.get("department", "")
    role = request.args.get("role", "")
    status = request.args.get("status", "")
    user_id = current_user_id()

    sql = "SELECT * FROM persons WHERE user_id = ?"
    params = [user_id]
    if q:
        sql += " AND (full_name LIKE ? OR person_code LIKE ? OR email LIKE ? OR phone LIKE ?)"
        like = f"%{q}%"
        params += [like, like, like, like]
    if department:
        sql += " AND department = ?"
        params.append(department)
    if role:
        sql += " AND role = ?"
        params.append(role)
    if status:
        sql += " AND status = ?"
        params.append(status)
    sql += " ORDER BY created_at DESC"

    conn = db.get_db()
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return jsonify({"success": True, "people": rows})


@app.route("/api/people/<int:person_id>", methods=["GET", "PUT", "DELETE"])
@login_required
def api_person_detail(person_id):
    conn = db.get_db()

    if request.method == "GET":
        person = get_owned_person(conn, person_id)
        if not person:
            conn.close()
            return jsonify({"success": False, "error": "Not found"}), 404

        total_visits = conn.execute(
            "SELECT COUNT(*) c FROM recognition_history WHERE person_id = ? AND is_known = 1", (person_id,)
        ).fetchone()["c"]
        present_days = conn.execute(
            "SELECT COUNT(*) c FROM attendance WHERE person_id = ?", (person_id,)
        ).fetchone()["c"]
        recent_history = [dict(r) for r in conn.execute(
            "SELECT * FROM recognition_history WHERE person_id = ? ORDER BY recognized_at DESC LIMIT 15",
            (person_id,)
        ).fetchall()]
        attendance_rows = [dict(r) for r in conn.execute(
            "SELECT * FROM attendance WHERE person_id = ? ORDER BY date DESC LIMIT 30", (person_id,)
        ).fetchall()]
        conn.close()
        return jsonify({
            "success": True,
            "person": dict(person),
            "total_visits": total_visits,
            "present_days": present_days,
            "recent_history": recent_history,
            "attendance": attendance_rows,
        })

    if request.method == "PUT":
        data = request.get_json(force=True)
        fields = ["full_name", "person_code", "age", "gender", "phone", "email",
                  "department", "role", "address", "dob", "status"]
        updates, params = [], []
        for f in fields:
            if f in data:
                updates.append(f"{f} = ?")
                params.append(data[f])
        if not updates:
            conn.close()
            return jsonify({"success": False, "error": "No fields to update"}), 400
        params.append(person_id)
        conn.execute(f"UPDATE persons SET {', '.join(updates)} WHERE id = ? AND user_id = ?", params + [current_user_id()])
        conn.commit()
        conn.close()
        return jsonify({"success": True})

    if request.method == "DELETE":
        person = get_owned_person(conn, person_id)
        if person and person["photo_path"]:
            full_path = os.path.join(BASE_DIR, person["photo_path"])
            if os.path.exists(full_path):
                try:
                    os.remove(full_path)
                except OSError:
                    pass
        conn.execute("DELETE FROM persons WHERE id = ? AND user_id = ?", (person_id, current_user_id()))  # cascades to encodings/attendance
        conn.commit()
        conn.close()
        return jsonify({"success": True})


@app.route("/api/people/<int:person_id>/face-data", methods=["DELETE"])
@login_required
def api_delete_face_data(person_id):
    conn = db.get_db()
    person = get_owned_person(conn, person_id)
    if not person:
        conn.close()
        return jsonify({"success": False, "error": "Not found"}), 404
    conn.execute("DELETE FROM face_encodings WHERE person_id = ?", (person_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Face data deleted. Person record kept."})


@app.route("/api/departments")
@login_required
def api_departments():
    conn = db.get_db()
    user_id = current_user_id()
    depts = [r["department"] for r in conn.execute(
        "SELECT DISTINCT department FROM persons WHERE user_id = ? AND department IS NOT NULL AND department != ''",
        (user_id,)
    ).fetchall()]
    roles = [r["role"] for r in conn.execute(
        "SELECT DISTINCT role FROM persons WHERE user_id = ? AND role IS NOT NULL AND role != ''",
        (user_id,)
    ).fetchall()]
    conn.close()
    return jsonify({"success": True, "departments": depts, "roles": roles})


# ---------------------------------------------------------------------------
# API — Attendance
# ---------------------------------------------------------------------------
@app.route("/api/attendance")
@login_required
def api_attendance():
    date_filter = request.args.get("date", "")
    department = request.args.get("department", "")
    q = request.args.get("q", "")

    user_id = current_user_id()
    sql = """
        SELECT a.*, p.full_name, p.person_code, p.department
        FROM attendance a JOIN persons p ON p.id = a.person_id
        WHERE p.user_id = ?
    """
    params = [user_id]
    if date_filter:
        sql += " AND a.date = ?"
        params.append(date_filter)
    if department:
        sql += " AND p.department = ?"
        params.append(department)
    if q:
        sql += " AND (p.full_name LIKE ? OR p.person_code LIKE ?)"
        params += [f"%{q}%", f"%{q}%"]
    sql += " ORDER BY a.date DESC, a.entry_time DESC"

    conn = db.get_db()
    rows = []
    for r in conn.execute(sql, params).fetchall():
        d = dict(r)
        d["duration"] = _calc_duration(d.get("entry_time"), d.get("exit_time"))
        rows.append(d)
    conn.close()
    return jsonify({"success": True, "attendance": rows})


def _calc_duration(entry, exit_):
    if not entry or not exit_:
        return None
    try:
        fmt = "%H:%M:%S"
        t1 = datetime.strptime(entry, fmt)
        t2 = datetime.strptime(exit_, fmt)
        delta = t2 - t1
        if delta.total_seconds() < 0:
            return None
        hours, remainder = divmod(int(delta.total_seconds()), 3600)
        minutes = remainder // 60
        return f"{hours}h {minutes}m"
    except Exception:
        return None


# ---------------------------------------------------------------------------
# API — Recognition history
# ---------------------------------------------------------------------------
@app.route("/api/history")
@login_required
def api_history():
    known_filter = request.args.get("known", "")  # "known" | "unknown" | ""
    date_filter = request.args.get("date", "")
    person = request.args.get("person", "")
    user_id = current_user_id()

    sql = "SELECT * FROM recognition_history WHERE user_id = ?"
    params = [user_id]
    if known_filter == "known":
        sql += " AND is_known = 1"
    elif known_filter == "unknown":
        sql += " AND is_known = 0"
    if date_filter:
        sql += " AND date(recognized_at) = ?"
        params.append(date_filter)
    if person:
        sql += " AND person_name LIKE ?"
        params.append(f"%{person}%")
    sql += " ORDER BY recognized_at DESC LIMIT 300"

    conn = db.get_db()
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return jsonify({"success": True, "history": rows})


# ---------------------------------------------------------------------------
# API — Unknown faces
# ---------------------------------------------------------------------------
@app.route("/api/unknown-faces")
@login_required
def api_unknown_faces():
    date_filter = request.args.get("date", "")
    user_id = current_user_id()
    sql = "SELECT id, snapshot_path, confidence, detection_count, first_seen, last_seen, resolved FROM unknown_faces WHERE user_id = ? AND resolved = 0"
    params = [user_id]
    if date_filter:
        sql += " AND date(first_seen) = ?"
        params.append(date_filter)
    sql += " ORDER BY last_seen DESC"

    conn = db.get_db()
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return jsonify({"success": True, "unknown_faces": rows})


@app.route("/api/unknown-faces/<int:unknown_id>", methods=["DELETE"])
@login_required
def api_delete_unknown(unknown_id):
    conn = db.get_db()
    row = conn.execute("SELECT * FROM unknown_faces WHERE id = ? AND user_id = ?", (unknown_id, current_user_id())).fetchone()
    if row and row["snapshot_path"]:
        full_path = os.path.join(BASE_DIR, row["snapshot_path"])
        if os.path.exists(full_path):
            try:
                os.remove(full_path)
            except OSError:
                pass
    conn.execute("DELETE FROM unknown_faces WHERE id = ? AND user_id = ?", (unknown_id, current_user_id()))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/unknown-faces/clear-all", methods=["DELETE"])
@login_required
def api_clear_unknown():
    conn = db.get_db()
    rows = conn.execute(
        "SELECT snapshot_path FROM unknown_faces WHERE user_id = ?",
        (current_user_id(),)
    ).fetchall()
    for r in rows:
        if r["snapshot_path"]:
            full_path = os.path.join(BASE_DIR, r["snapshot_path"])
            if os.path.exists(full_path):
                try:
                    os.remove(full_path)
                except OSError:
                    pass
    conn.execute("DELETE FROM unknown_faces WHERE user_id = ?", (current_user_id(),))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# API — Reports
# ---------------------------------------------------------------------------
@app.route("/api/reports/summary")
@login_required
def api_reports_summary():
    conn = db.get_db()
    user_id = current_user_id()

    # last 7 days attendance trend
    trend = []
    for i in range(6, -1, -1):
        d = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
        c = conn.execute(
            "SELECT COUNT(*) c FROM attendance a JOIN persons p ON p.id = a.person_id "
            "WHERE p.user_id = ? AND a.date = ?", (user_id, d)
        ).fetchone()["c"]
        trend.append({"date": d, "count": c})

    known_count = conn.execute(
        "SELECT COUNT(*) c FROM recognition_history WHERE user_id = ? AND is_known = 1", (user_id,)
    ).fetchone()["c"]
    unknown_count = conn.execute(
        "SELECT COUNT(*) c FROM recognition_history WHERE user_id = ? AND is_known = 0", (user_id,)
    ).fetchone()["c"]

    dept_attendance = [dict(r) for r in conn.execute("""
        SELECT p.department as department, COUNT(*) as count
        FROM attendance a JOIN persons p ON p.id = a.person_id
        WHERE p.user_id = ? AND p.department IS NOT NULL AND p.department != ''
        GROUP BY p.department
    """, (user_id,)).fetchall()]

    top_people = [dict(r) for r in conn.execute("""
        SELECT p.full_name, COUNT(*) as visits
        FROM recognition_history rh JOIN persons p ON p.id = rh.person_id
        WHERE rh.user_id = ? AND rh.is_known = 1 AND p.user_id = ?
        GROUP BY rh.person_id ORDER BY visits DESC LIMIT 5
    """, (user_id, user_id)).fetchall()]

    activity_by_hour = [dict(r) for r in conn.execute("""
        SELECT strftime('%H', recognized_at) as hour, COUNT(*) as count
        FROM recognition_history
        WHERE user_id = ? AND date(recognized_at) = date('now')
        GROUP BY hour ORDER BY hour
    """, (user_id,)).fetchall()]

    conn.close()
    return jsonify({
        "success": True,
        "attendance_trend": trend,
        "known_vs_unknown": {"known": known_count, "unknown": unknown_count},
        "department_attendance": dept_attendance,
        "top_people": top_people,
        "activity_by_hour": activity_by_hour,
    })


@app.route("/api/reports/export/<fmt>")
@login_required
def api_reports_export(fmt):
    conn = db.get_db()
    user_id = current_user_id()
    rows = conn.execute("""
        SELECT p.full_name, p.person_code, p.department, a.date, a.entry_time, a.exit_time, a.status
        FROM attendance a JOIN persons p ON p.id = a.person_id
        WHERE p.user_id = ?
        ORDER BY a.date DESC
    """, (user_id,)).fetchall()
    conn.close()

    headers = ["Full Name", "ID", "Department", "Date", "Entry Time", "Exit Time", "Status"]

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        for r in rows:
            writer.writerow([r["full_name"], r["person_code"], r["department"], r["date"], r["entry_time"], r["exit_time"], r["status"]])
        mem = io.BytesIO(buf.getvalue().encode("utf-8"))
        return send_file(mem, mimetype="text/csv", as_attachment=True, download_name="attendance_report.csv")

    elif fmt == "excel":
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({"success": False, "error": "openpyxl not installed. Run: pip install openpyxl"}), 503

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(headers)
        for r in rows:
            ws.append([r["full_name"], r["person_code"], r["department"], r["date"], r["entry_time"], r["exit_time"], r["status"]])
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 2)

        mem = io.BytesIO()
        wb.save(mem)
        mem.seek(0)
        return send_file(mem, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          as_attachment=True, download_name="attendance_report.xlsx")

    elif fmt == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            return jsonify({"success": False, "error": "reportlab not installed. Run: pip install reportlab"}), 503

        mem = io.BytesIO()
        doc = SimpleDocTemplate(mem, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = [Paragraph("Attendance Report", styles["Title"])]

        table_data = [headers] + [
            [r["full_name"], r["person_code"] or "-", r["department"] or "-", r["date"], r["entry_time"] or "-", r["exit_time"] or "-", r["status"]]
            for r in rows
        ]
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6C5CE7")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F6FF")]),
        ]))
        elements.append(table)
        doc.build(elements)
        mem.seek(0)
        return send_file(mem, mimetype="application/pdf", as_attachment=True, download_name="attendance_report.pdf")

    return jsonify({"success": False, "error": "Unknown format"}), 400


# ---------------------------------------------------------------------------
# API — Settings
# ---------------------------------------------------------------------------
@app.route("/api/settings", methods=["GET", "POST"])
@login_required
def api_settings():
    if request.method == "GET":
        conn = db.get_db()
        rows = conn.execute("SELECT key, value FROM settings").fetchall()
        conn.close()
        return jsonify({"success": True, "settings": {r["key"]: r["value"] for r in rows}})

    data = request.get_json(force=True)
    allowed_keys = {
        "recognition_threshold", "detection_sensitivity", "camera_device",
        "resolution", "fps", "session_timeout", "unknown_cooldown_seconds"
    }
    for k, v in data.items():
        if k in allowed_keys:
            db.set_setting(k, v)
    return jsonify({"success": True})


@app.route("/api/settings/change-password", methods=["POST"])
@login_required
def api_change_password():
    data = request.get_json(force=True)
    current = data.get("current_password", "")
    new = data.get("new_password", "")

    if not new or len(new) < 6:
        return jsonify({"success": False, "error": "New password must be at least 6 characters"}), 400

    conn = db.get_db()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (session["user_id"],)).fetchone()
    if not user or not check_password_hash(user["password_hash"], current):
        conn.close()
        return jsonify({"success": False, "error": "Current password is incorrect"}), 401

    conn.execute("UPDATE users SET password_hash = ? WHERE id = ?",
                 (generate_password_hash(new), session["user_id"]))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/settings/backup-database")
@login_required
def api_backup_database():
    return send_file(db.DB_PATH, as_attachment=True,
                      download_name=f"database_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")


@app.route("/api/settings/delete-all-unknown", methods=["DELETE"])
@login_required
def api_delete_all_unknown():
    return api_clear_unknown()


@app.route("/api/settings/delete-all-face-data", methods=["DELETE"])
@login_required
def api_delete_all_face_data():
    conn = db.get_db()
    conn.execute(
        "DELETE FROM face_encodings WHERE person_id IN (SELECT id FROM persons WHERE user_id = ?)",
        (current_user_id(),)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "All stored face encodings deleted for your account. Person profiles kept."})


# ---------------------------------------------------------------------------
# Error handlers
# ---------------------------------------------------------------------------
@app.errorhandler(404)
def not_found(e):
    if request.path.startswith("/api/"):
        return jsonify({"success": False, "error": "Not found"}), 404
    return render_template("login.html"), 404


@app.errorhandler(500)
def server_error(e):
    if request.path.startswith("/api/"):
        return jsonify({"success": False, "error": str(e) or "Internal server error. Please try again."}), 500
    return render_template("login.html"), 500


if __name__ == "__main__":
    db.init_db()
    print("=" * 60)
    print(" AI Face Recognition & Smart Attendance System")
    print(" Running at: http://127.0.0.1:5000")
    print(" Demo login: admin / admin123")
    if not fu.FACE_RECOGNITION_AVAILABLE:
        print(" WARNING: 'face_recognition' library not found.")
        print(" Face detection/recognition features are disabled until installed.")
    print("=" * 60)
    app.run(debug=True, host="0.0.0.0", port=5000)
